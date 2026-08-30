import io
from typing import List, Optional, Union
from pydantic import BaseModel, Field
from fastapi import APIRouter, Query, HTTPException, Body
from fastapi.responses import StreamingResponse
import pandas as pd

from . import geocode, overpass_client, cleaning

router = APIRouter()

class SearchRequest(BaseModel):
    country: str = Field(..., description="Country name (required)")
    city: Optional[str] = Field(None, description="City name (optional)")
    category: Optional[str] = Field(None, description="Category or tag filter")
    is_custom_category: bool = Field(False, description="True if category is a raw OSM tag string")
    limit_mode: str = Field("capped", description="'capped', 'increased', or 'disabled'")
    limit_value: int = Field(500, description="Max POIs to fetch if limit_mode is capped or increased")
    output_mode: str = Field("session", description="'session' for JSON response or 'export'/'excel' for file download")
    format: str = Field("json", description="'json' or 'excel'")
    filters: Optional[Union[List[str], str]] = Field(None, description="Optional raw OSM tag filter list or string")

def _generate_excel_response(df: pd.DataFrame, country: str, city: Optional[str], category: Optional[str]) -> StreamingResponse:
    """Generate an Excel workbook from the cleaned DataFrame with openpyxl styling."""
    excel_cols = ["name", "category", "phone", "website", "address", "lat", "lon"]
    # Ensure columns exist, fill if missing
    for col in excel_cols:
        if col not in df.columns:
            df[col] = ""
            
    df_export = df[excel_cols].copy()
    df_export.columns = ["Name", "Category", "Phone", "Website", "Address", "Latitude", "Longitude"]
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    
    # Sanitize and truncate sheet title (max 31 characters, no invalid chars)
    sheet_title = f"{country}_{city or ''}_{category or ''}"
    for char in r"\*?/:[]":
        sheet_title = sheet_title.replace(char, "")
    ws.title = sheet_title[:31] if sheet_title else "Leads"
    
    # Header styling
    headers = list(df_export.columns)
    font_bold_white = Font(bold=True, color="FFFFFFFF")
    fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Dark steel blue
    align_center = Alignment(horizontal="center", vertical="center")
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = font_bold_white
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
        
    # Cell formatting and alternating fills
    fill_odd = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for row_idx, row_data in enumerate(df_export.itertuples(index=False), start=2):
        row_fill = fill_odd if row_idx % 2 == 1 else fill_even
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value="" if pd.isna(val) else val)
            cell.fill = row_fill
            cell.border = border_cell
            
            # Alignments
            col_name = headers[col_idx - 1]
            if col_name in ["Latitude", "Longitude"]:
                cell.alignment = Alignment(horizontal="right")
            elif col_name in ["Phone"]:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Freeze panes
    ws.freeze_panes = "A2"
    
    # AutoFilter
    if len(df_export) > 0:
        ws.auto_filter.ref = ws.dimensions
        
    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        fit_width = max(12, min(60, max_len + 3))
        ws.column_dimensions[col_letter].width = fit_width
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"leads_{country}_{city or ''}.xlsx".replace(" ", "_")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

def _execute_search(req: SearchRequest):
    """Core lead-generation pipeline execution."""
    try:
        bbox = geocode.geocode(req.country, req.city)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Geocoding failed: {str(exc)}")

    try:
        all_records, is_custom_match, is_chunked = overpass_client.query_pois(
            bbox=bbox,
            category=req.category,
            is_custom_category=req.is_custom_category,
            limit_mode=req.limit_mode,
            limit_value=req.limit_value,
        )
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Overpass query failed: {str(exc)}")

    total_found = len(all_records)
    df = cleaning.process_pois(all_records)
    total_after_dedup = len(df)

    is_excel = req.format.lower() == "excel" or req.output_mode.lower() in ("export", "excel")

    if is_excel:
        return _generate_excel_response(df, req.country, req.city, req.category)
    else:
        results = cleaning.prepare_json_records(df)
        return {
            "status": "success",
            "country": req.country,
            "city": req.city,
            "total_found": total_found,
            "total_after_dedup": total_after_dedup,
            "custom_fuzzy_used": is_custom_match,
            "chunked": is_chunked,
            "count": len(results),
            "results": results,
        }

@router.post("/search")
def search_post(req: SearchRequest = Body(...)):
    """POST /search endpoint for JSON requests."""
    return _execute_search(req)

@router.get("/search")
def search_get(
    country: str = Query(..., description="Country name (required)"),
    city: Optional[str] = Query(None, description="City name (optional)"),
    category: Optional[str] = Query(None, description="Category name"),
    is_custom_category: bool = Query(False, description="Is custom category tag string"),
    limit_mode: str = Query("capped", description="'capped', 'increased', or 'disabled'"),
    limit_value: int = Query(500, description="Limit count"),
    output_mode: str = Query("session", description="'session' or 'excel'/'export'"),
    format: str = Query("json", pattern="^(json|excel)$", description="Response format: json or excel"),
):
    """GET /search endpoint for query parameter requests."""
    req = SearchRequest(
        country=country,
        city=city,
        category=category,
        is_custom_category=is_custom_category,
        limit_mode=limit_mode,
        limit_value=limit_value,
        output_mode=output_mode,
        format=format
    )
    return _execute_search(req)
